"""BESS Sizing Tool — Excel Export Module

Generates an Excel report with sizing results using openpyxl.
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# LG brand colors
LG_RED = "A50034"
LG_RED_LIGHT = "D4003D"
LG_GRAY = "4A4A4A"
LG_LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"

# --- Style helpers ---

def _header_font(size=11):
    return Font(bold=True, size=size, color=WHITE)


def _label_font(bold=False):
    return Font(bold=bold, size=10, color=LG_GRAY)


def _header_fill():
    return PatternFill(start_color=LG_RED, end_color=LG_RED, fill_type="solid")


def _subheader_fill():
    return PatternFill(start_color=LG_GRAY, end_color=LG_GRAY, fill_type="solid")


def _alt_fill():
    return PatternFill(start_color=LG_LIGHT_GRAY, end_color=LG_LIGHT_GRAY, fill_type="solid")


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left():
    return Alignment(horizontal="left", vertical="center")


def _right():
    return Alignment(horizontal="right", vertical="center")


def _apply_header_row(ws, row, cols, label):
    """Merge cols and apply LG red header to a single label cell."""
    ws.merge_cells(start_row=row, start_column=cols[0],
                   end_row=row, end_column=cols[-1])
    cell = ws.cell(row=row, column=cols[0])
    cell.value = label
    cell.font = _header_font(11)
    cell.fill = _header_fill()
    cell.alignment = _center()
    cell.border = _thin_border()


def _write_row(ws, row, pairs, alt=False):
    """Write (label, value) pairs across two-column groups starting at col 1."""
    col = 1
    for label, value in pairs:
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = _label_font(bold=True)
        lc.border = _thin_border()
        lc.alignment = _left()
        if alt:
            lc.fill = _alt_fill()

        vc = ws.cell(row=row, column=col + 1, value=value)
        vc.font = _label_font()
        vc.border = _thin_border()
        vc.alignment = _right()
        if alt:
            vc.fill = _alt_fill()
        col += 2


def _col_widths(ws, widths: dict):
    """Set column widths. widths is {col_letter: width}."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# --- Sheet builders ---

def _build_summary(ws, result: dict):
    inp = result.get("input_data", {})
    battery = result.get("battery", {})
    reactive = result.get("reactive_power", {})
    efficiency = result.get("efficiency", {})

    # Title block
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = "BESS SIZING REPORT"
    title_cell.font = Font(bold=True, size=16, color=WHITE)
    title_cell.fill = _header_fill()
    title_cell.alignment = _center()

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    # Project info block (rows 4-10)
    _apply_header_row(ws, 4, list(range(1, 9)), "PROJECT INFORMATION")
    info_pairs = [
        ("Project Title", inp.get("project_title", "—")),
        ("Customer", inp.get("customer", "—")),
        ("Date", str(date.today())),
        ("Application", inp.get("application", "—")),
        ("POI Voltage", inp.get("poi_voltage", "—")),
        ("Temperature (°C)", inp.get("temperature_c", "—")),
        ("Project Life (yr)", inp.get("project_life_yr", "—")),
        ("Scope of Supply", inp.get("scope_of_supply", "—")),
    ]
    for i, (label, value) in enumerate(info_pairs):
        row = 5 + (i // 2)
        col = 1 if i % 2 == 0 else 5
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = _label_font(bold=True)
        lc.border = _thin_border()
        lc.alignment = _left()
        if i % 2 == 1:
            lc.fill = _alt_fill()
        vc = ws.cell(row=row, column=col + 1, value=value)
        vc.font = _label_font()
        vc.border = _thin_border()
        vc.alignment = _right()
        if i % 2 == 1:
            vc.fill = _alt_fill()

    # Key Metrics (rows 12+)
    _apply_header_row(ws, 12, list(range(1, 9)), "KEY METRICS")
    metrics = [
        ("Required Power @POI (MW)", inp.get("required_power_mw", "—")),
        ("Required Energy @POI (MWh)", inp.get("required_energy_mwh", "—")),
        ("Installed Energy DC (MWh)", round(battery.get("installation_energy_dc_mwh", 0), 2)),
        ("Usable Energy @POI (MWh)", round(battery.get("dischargeable_energy_poi_mwh", 0), 2)),
        ("Duration @BOL (hr)", round(battery.get("duration_bol_hr", 0), 2)),
        ("CP Rate (DC)", round(battery.get("cp_rate", 0), 4)),
        ("No. of PCS", battery.get("no_of_pcs", "—")),
        ("No. of LINKs", battery.get("no_of_links", "—")),
        ("No. of Racks", battery.get("no_of_racks", "—")),
        ("No. of MVT", battery.get("no_of_mvt", "—")),
        ("System Efficiency (%)", round(efficiency.get("total_bat_poi_eff", 0) * 100, 2) if efficiency else "—"),
        ("Power Factor", inp.get("power_factor", "—")),
        ("HV Apparent Power (kVA)", round(reactive.get("total_apparent_power_poi_kva", 0), 1) if reactive else "—"),
        ("PCS Capacity OK", "YES" if reactive and reactive.get("is_pcs_sufficient") else "NO"),
    ]
    for i, (label, value) in enumerate(metrics):
        row = 13 + (i // 2)
        col = 1 if i % 2 == 0 else 5
        alt = (i // 2) % 2 == 1
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = _label_font(bold=True)
        lc.border = _thin_border()
        lc.alignment = _left()
        if alt:
            lc.fill = _alt_fill()
        vc = ws.cell(row=row, column=col + 1, value=value)
        vc.font = _label_font()
        vc.border = _thin_border()
        vc.alignment = _right()
        if alt:
            vc.fill = _alt_fill()

    _col_widths(ws, {
        "A": 30, "B": 18, "C": 4, "D": 4,
        "E": 30, "F": 18, "G": 4, "H": 4,
    })
    _freeze(ws, "A3")


def _build_battery_sizing(ws, result: dict):
    battery = result.get("battery", {})
    inp = result.get("input_data", {})
    pcs = result.get("pcs", {})

    _apply_header_row(ws, 1, [1, 2, 3], "BATTERY SIZING DETAIL")

    sections = [
        ("POI Requirements", [
            ("Required Power @POI", f"{inp.get('required_power_mw', '—')} MW"),
            ("Required Energy @POI", f"{inp.get('required_energy_mwh', '—')} MWh"),
            ("Power Factor", inp.get("power_factor", "—")),
        ]),
        ("DC Requirements", [
            ("Required Power @DC (MW)", round(battery.get("req_power_dc_mw", 0), 3)),
            ("Required Energy @DC (MWh)", round(battery.get("req_energy_dc_mwh", 0), 3)),
        ]),
        ("PCS Configuration", [
            ("PCS Type", inp.get("pcs_type", "—")),
            ("PCS Unit Power (MW)", round(pcs.get("pcs_unit_power_mw", 0), 5) if pcs else "—"),
            ("No. of PCS", battery.get("no_of_pcs", "—")),
            ("No. of MVT", battery.get("no_of_mvt", "—")),
        ]),
        ("LINK / Rack Configuration", [
            ("Product Type", inp.get("product_type_a", "—")),
            ("No. of LINKs", battery.get("no_of_links", "—")),
            ("Racks per LINK", battery.get("racks_per_link", "—")),
            ("Total Racks", battery.get("no_of_racks", "—")),
            ("Nameplate per LINK (MWh)", battery.get("nameplate_energy_per_link_mwh", "—")),
        ]),
        ("Installation Energy", [
            ("Installation Energy DC (MWh)", round(battery.get("installation_energy_dc_mwh", 0), 3)),
            ("CP Rate", round(battery.get("cp_rate", 0), 6)),
            ("Dischargeable Energy DC (MWh)", round(battery.get("dischargeable_energy_dc_mwh", 0), 3)),
            ("Dischargeable Energy @POI (MWh)", round(battery.get("dischargeable_energy_poi_mwh", 0), 3)),
            ("Duration @BOL (hr)", round(battery.get("duration_bol_hr", 0), 3)),
            ("Aux Power Peak (MW)", round(battery.get("aux_power_peak_mw", 0), 3)),
        ]),
    ]

    row = 2
    for section_title, items in sections:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        hc = ws.cell(row=row, column=1, value=section_title)
        hc.font = Font(bold=True, size=10, color=WHITE)
        hc.fill = _subheader_fill()
        hc.alignment = _left()
        hc.border = _thin_border()
        row += 1

        for i, (label, value) in enumerate(items):
            alt = i % 2 == 1
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = _label_font(bold=True)
            lc.border = _thin_border()
            lc.alignment = _left()
            if alt:
                lc.fill = _alt_fill()
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = _label_font()
            vc.border = _thin_border()
            vc.alignment = _right()
            if alt:
                vc.fill = _alt_fill()
            row += 1

        row += 1  # blank separator

    _col_widths(ws, {"A": 35, "B": 20, "C": 10})
    _freeze(ws, "A2")


def _build_retention(ws, result: dict):
    retention = result.get("retention", {})
    by_year = retention.get("retention_by_year", {}) if retention else {}
    battery = result.get("battery", {})

    _apply_header_row(ws, 1, [1, 2, 3, 4], "RETENTION CURVE")

    # Table headers
    headers = ["Year", "Retention (%)", "Total Energy (MWh)", "Dischargeable @POI (MWh)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = _subheader_fill()
        c.alignment = _center()
        c.border = _thin_border()

    data_start = 3
    for i, (year_key, yr_data) in enumerate(sorted(by_year.items(), key=lambda x: int(x[0]))):
        year = int(year_key)
        ret_pct = yr_data.get("retention_pct", 0)
        total_energy = yr_data.get("total_energy_mwh", 0)
        dischargeable_poi = yr_data.get("dischargeable_energy_poi_mwh", 0)

        alt = i % 2 == 1
        row = data_start + i
        vals = [year, ret_pct, total_energy, dischargeable_poi]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = _label_font()
            c.border = _thin_border()
            c.alignment = _center()
            if alt:
                c.fill = _alt_fill()

    # Line chart
    num_rows = len(by_year)
    if num_rows > 0:
        chart = LineChart()
        chart.title = "Retention Curve"
        chart.style = 10
        chart.y_axis.title = "Energy (MWh)"
        chart.x_axis.title = "Year"
        chart.height = 12
        chart.width = 20

        # Total energy series
        data_ref = Reference(ws, min_col=3, min_row=2,
                             max_col=4, max_row=data_start + num_rows - 1)
        cats = Reference(ws, min_col=1, min_row=data_start,
                         max_row=data_start + num_rows - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.solidFill = LG_RED
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.line.solidFill = LG_GRAY

        ws.add_chart(chart, f"F2")

    _col_widths(ws, {"A": 8, "B": 16, "C": 22, "D": 26})
    _freeze(ws, "A3")


def _build_reactive_power(ws, result: dict):
    reactive = result.get("reactive_power", {})
    inp = result.get("input_data", {})
    battery = result.get("battery", {})
    pcs = result.get("pcs", {})

    _apply_header_row(ws, 1, [1, 2, 3], "REACTIVE POWER ANALYSIS")

    sections = [
        ("HV Level", [
            ("Required Power @POI (MW)", inp.get("required_power_mw", "—")),
            ("Power Factor", inp.get("power_factor", "—")),
            ("Total Apparent Power @POI (kVA)", round(reactive.get("total_apparent_power_poi_kva", 0), 3)),
            ("Grid Reactive Power (kVAR)", round(reactive.get("grid_kvar", 0), 3)),
            ("HV Transformer Reactive (kVAR)", round(reactive.get("hv_tr_kvar", 0), 3)),
            ("HV P-Loss (kW)", round(reactive.get("p_loss_hv_kw", 0), 3)),
        ]),
        ("MV Level", [
            ("Power Factor @MV", round(reactive.get("pf_at_mv", 0), 6)),
        ]),
        ("Inverter Level", [
            ("Total Required S (kVA)", round(reactive.get("total_s_inverter_kva", 0), 3)),
            ("Available S Total (kVA)", reactive.get("available_s_total_kva", "—")),
            ("PCS Capacity Sufficient", "YES" if reactive.get("is_pcs_sufficient") else "NO"),
        ]),
        ("PCS Capacity Check", [
            ("No. of PCS", battery.get("no_of_pcs", "—")),
            ("PCS Unit kVA", pcs.get("pcs_unit_kva", "—") if pcs else "—"),
            ("Total Nameplate kVA", reactive.get("available_s_total_kva", "—")),
            ("Required kVA", round(reactive.get("total_s_inverter_kva", 0), 1)),
            ("Margin (kVA)", round(
                reactive.get("available_s_total_kva", 0) - reactive.get("total_s_inverter_kva", 0), 1
            ) if reactive else "—"),
        ]),
    ]

    row = 2
    for section_title, items in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        hc = ws.cell(row=row, column=1, value=section_title)
        hc.font = Font(bold=True, size=10, color=WHITE)
        hc.fill = _subheader_fill()
        hc.alignment = _left()
        hc.border = _thin_border()
        row += 1

        for i, (label, value) in enumerate(items):
            alt = i % 2 == 1
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = _label_font(bold=True)
            lc.border = _thin_border()
            lc.alignment = _left()
            if alt:
                lc.fill = _alt_fill()
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = _label_font()
            vc.border = _thin_border()
            vc.alignment = _right()
            if alt:
                vc.fill = _alt_fill()
            row += 1

        row += 1

    _col_widths(ws, {"A": 35, "B": 20, "C": 10})
    _freeze(ws, "A2")


def _build_efficiency(ws, result: dict):
    efficiency = result.get("efficiency", {})
    inp = result.get("input_data", {})
    eff_inp = inp.get("efficiency", {}) if inp else {}
    bat_loss = inp.get("battery_loss", {}) if inp else {}
    aux_eff = inp.get("aux_efficiency", {}) if inp else {}

    _apply_header_row(ws, 1, [1, 2, 3], "EFFICIENCY CHAIN")

    sections = [
        ("System Efficiency Chain", [
            ("HV AC Cabling", f"{round(eff_inp.get('hv_ac_cabling', 0) * 100, 2)} %"),
            ("HV Transformer", f"{round(eff_inp.get('hv_transformer', 0) * 100, 2)} %"),
            ("MV AC Cabling", f"{round(eff_inp.get('mv_ac_cabling', 0) * 100, 2)} %"),
            ("MV Transformer", f"{round(eff_inp.get('mv_transformer', 0) * 100, 2)} %"),
            ("LV Cable", f"{round(eff_inp.get('lv_cabling', 0) * 100, 2)} %"),
            ("PCS Efficiency", f"{round(eff_inp.get('pcs_efficiency', 0) * 100, 2)} %"),
            ("DC Cabling", f"{round(eff_inp.get('dc_cabling', 0) * 100, 2)} %"),
            ("Total Bat→POI Efficiency", f"{round(efficiency.get('total_bat_poi_eff', 0) * 100, 4)} %"),
        ]),
        ("Auxiliary Efficiency", [
            ("Aux Branching Point", aux_eff.get("branching_point", "—")),
            ("Total Aux Eff (MV)", f"{round(aux_eff.get('total_aux_eff_mv', 0) * 100, 4)} %"),
            ("Total DC→Aux Eff", f"{round(aux_eff.get('total_dc_to_aux_eff', 0) * 100, 4)} %"),
        ]),
        ("Battery Loss Factors", [
            ("Applied DOD", f"{round(bat_loss.get('applied_dod', 0) * 100, 2)} %"),
            ("Loss Factors", round(bat_loss.get("loss_factors", 0), 5)),
            ("MBMS Consumption", round(bat_loss.get("mbms_consumption", 0), 4)),
            ("Total Battery Loss Factor", round(efficiency.get("total_battery_loss_factor", 0), 6)),
        ]),
        ("Overall", [
            ("Total System Efficiency", f"{round(efficiency.get('total_efficiency', 0) * 100, 4)} %" if efficiency else "—"),
        ]),
    ]

    row = 2
    for section_title, items in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        hc = ws.cell(row=row, column=1, value=section_title)
        hc.font = Font(bold=True, size=10, color=WHITE)
        hc.fill = _subheader_fill()
        hc.alignment = _left()
        hc.border = _thin_border()
        row += 1

        for i, (label, value) in enumerate(items):
            alt = i % 2 == 1
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = _label_font(bold=True)
            lc.border = _thin_border()
            lc.alignment = _left()
            if alt:
                lc.fill = _alt_fill()
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = _label_font()
            vc.border = _thin_border()
            vc.alignment = _right()
            if alt:
                vc.fill = _alt_fill()
            row += 1

        row += 1

    _col_widths(ws, {"A": 35, "B": 22, "C": 10})
    _freeze(ws, "A2")


# --- Public API ---

def generate_excel_report(result: dict) -> io.BytesIO:
    """Generate Excel report from calculation results.

    Args:
        result: Dict with keys: efficiency, pcs, battery, retention,
                reactive_power, rte, input_data

    Returns:
        BytesIO buffer containing the Excel file.
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary(ws_summary, result)

    # Sheet 2: Battery Sizing Detail
    ws_sizing = wb.create_sheet("Battery Sizing")
    _build_battery_sizing(ws_sizing, result)

    # Sheet 3: Retention Curve
    ws_retention = wb.create_sheet("Retention")
    _build_retention(ws_retention, result)

    # Sheet 4: Reactive Power
    ws_reactive = wb.create_sheet("Reactive Power")
    _build_reactive_power(ws_reactive, result)

    # Sheet 5: Efficiency Chain
    ws_eff = wb.create_sheet("Efficiency")
    _build_efficiency(ws_eff, result)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- Comparison Excel ---

def generate_comparison_excel(cases: list) -> io.BytesIO:
    """Generate a comparison Excel report for multiple cases.

    Args:
        cases: List of case dicts, each with case_name, input_data, result_data.

    Returns:
        BytesIO buffer containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Title block
    num_cases = len(cases)
    total_cols = num_cases + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value="BESS CASE COMPARISON")
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = _header_fill()
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Header row: Metric | Case1 | Case2 | ...
    hdr = ws.cell(row=3, column=1, value="Metric")
    hdr.font = _header_font(10)
    hdr.fill = _subheader_fill()
    hdr.border = _thin_border()
    hdr.alignment = _left()
    for ci, case in enumerate(cases):
        c = ws.cell(row=3, column=ci + 2, value=case.get('case_name', f'Case {ci+1}'))
        c.font = _header_font(10)
        c.fill = _subheader_fill()
        c.border = _thin_border()
        c.alignment = _center()

    # Metric definitions: (section, label, path, fmt)
    metrics = [
        ("System Configuration", None, None, None),
        (None, "Battery Product", "input.product_type", "text"),
        (None, "PCS Configuration", "input.pcs_type", "text"),
        (None, "Temperature (\u00b0C)", "input.temperature_c", "int"),
        ("Requirements", None, None, None),
        (None, "Required Power @POI (MW)", "input.required_power_mw", "num2"),
        (None, "Required Energy @POI (MWh)", "input.required_energy_mwh", "num2"),
        ("Sizing Results", None, None, None),
        (None, "No. of PCS", "result.battery.no_of_pcs", "int"),
        (None, "No. of LINKs", "result.battery.no_of_links", "int"),
        (None, "No. of Racks", "result.battery.no_of_racks", "int"),
        (None, "Installation Energy DC (MWh)", "result.battery.installation_energy_dc_mwh", "num2"),
        (None, "Dischargeable @POI (MWh)", "result.battery.dischargeable_energy_poi_mwh", "num2"),
        (None, "Duration @BOL (hr)", "result.battery.duration_bol_hr", "num2"),
        (None, "CP Rate", "result.battery.cp_rate", "num4"),
        ("Efficiency", None, None, None),
        (None, "Bat\u2192POI Efficiency (%)", "result.efficiency.total_bat_poi_eff", "pct"),
        (None, "System RTE (%)", "result.rte.system_rte", "pct"),
        ("Retention", None, None, None),
        (None, "Retention @Y10 (%)", "result.retention.retention_by_year.10.retention_pct", "num2"),
        (None, "Retention @Y20 (%)", "result.retention.retention_by_year.20.retention_pct", "num2"),
    ]

    def _get_nested(data, path):
        parts = path.split('.')
        obj = data
        if parts[0] == 'input':
            obj = data.get('input_data', {})
            parts = parts[1:]
        elif parts[0] == 'result':
            obj = data.get('result_data', {})
            parts = parts[1:]
        for p in parts:
            if obj is None:
                return None
            obj = obj.get(p) if isinstance(obj, dict) else None
        return obj

    def _fmt(v, fmt):
        if v is None:
            return "\u2014"
        if fmt == "text":
            return str(v)
        if fmt == "int":
            return int(round(float(v)))
        if fmt == "num2":
            return round(float(v), 2)
        if fmt == "num4":
            return round(float(v), 4)
        if fmt == "pct":
            return round(float(v) * 100, 2)
        return v

    row = 4
    for section, label, path, fmt in metrics:
        if section is not None:
            # Section header row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
            sc = ws.cell(row=row, column=1, value=section)
            sc.font = Font(bold=True, size=10, color=WHITE)
            sc.fill = PatternFill(start_color=LG_RED, end_color=LG_RED, fill_type="solid")
            sc.alignment = _left()
            sc.border = _thin_border()
            row += 1
            continue

        # Label cell
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _label_font(bold=True)
        lc.border = _thin_border()
        lc.alignment = _left()
        if (row - 4) % 2 == 0:
            lc.fill = _alt_fill()

        # Value cells
        for ci, case in enumerate(cases):
            raw = _get_nested(case, path)
            val = _fmt(raw, fmt)
            vc = ws.cell(row=row, column=ci + 2, value=val)
            vc.font = _label_font()
            vc.border = _thin_border()
            vc.alignment = _right()
            if (row - 4) % 2 == 0:
                vc.fill = _alt_fill()

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    for ci in range(num_cases):
        ws.column_dimensions[get_column_letter(ci + 2)].width = 22

    _freeze(ws, "A4")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
